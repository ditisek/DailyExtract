import wx
import codecs
import pandas as pd
import openpyxl
import psutil

p = psutil.Process()

# begin wxGlade: dependencies
# end wxGlade

# begin wxGlade: extracode
# end wxGlade


class MainFrame(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: MainFrame.__init__
        kwds["style"] = kwds.get("style", 0) | wx.CAPTION | \
                        wx.CLIP_CHILDREN | wx.CLOSE_BOX | wx.MAXIMIZE_BOX | \
                        wx.RESIZE_BORDER | wx.SYSTEM_MENU
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((930, 773))
        self.SetTitle("Extract Messages")

        self.window_1 = wx.SplitterWindow(self, wx.ID_ANY)
        # self.window_1.SetMinSize((628, 734))
        self.window_1.SetSize((380, 750))
        self.window_1.SetMinimumPaneSize(20)

        self.window_1_pane_1 = wx.Panel(self.window_1, wx.ID_ANY)
        # self.window_1_pane_1.SetMinSize((140, 734))
        self.window_1_pane_1.SetSize((240, 734))

        sizer_1 = wx.BoxSizer(wx.VERTICAL)

        sizer_3 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(sizer_3, 0, wx.EXPAND, 0)

        label_1 = wx.StaticText(self.window_1_pane_1, wx.ID_ANY,
                                "File extension: *.")
        sizer_3.Add(label_1, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)

        self.text_ctrl_1 = wx.TextCtrl(self.window_1_pane_1, wx.ID_ANY,
                                       "xlsm")
        sizer_3.Add(self.text_ctrl_1, 0, wx.ALIGN_CENTER_VERTICAL, 0)

        self.button_1 = wx.Button(self.window_1_pane_1, wx.ID_ANY,
                                  "Select files")
        sizer_1.Add(self.button_1, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        self.button_3 = wx.Button(self.window_1_pane_1, wx.ID_ANY,
                                  "Read headers from files")
        sizer_1.Add(self.button_3, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        self.check_list_box_2 = wx.CheckListBox(
            self.window_1_pane_1, wx.ID_ANY, choices=["headers"])
        self.check_list_box_2.SetMinSize((101, 200))
        sizer_1.Add(self.check_list_box_2, 0, wx.ALIGN_CENTER_HORIZONTAL, 0)

        self.button_2 = wx.Button(self.window_1_pane_1, wx.ID_ANY,
                                  "Export selected")
        sizer_1.Add(self.button_2, 0,
                    wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        self.window_1_pane_2 = wx.Panel(self.window_1, wx.ID_ANY)
        self.window_1_pane_2.SetSize((300, 750))

        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)

        self.check_list_box_1 = wx.CheckListBox(
            self.window_1_pane_2, wx.ID_ANY, choices=["choice 1"])
        # self.check_list_box_1.SetSize((300, 750))
        self.check_list_box_1.SetMinSize((650, 730))
        sizer_2.Add(self.check_list_box_1, 0, wx.ALL | wx.EXPAND, 0)
        # sizer_2.Add(self.check_list_box_1, 0, wx.ALL, 0)

        self.window_1_pane_2.SetSizer(sizer_2)

        self.window_1_pane_1.SetSizer(sizer_1)

        self.window_1.SplitVertically(self.window_1_pane_1,
                                      self.window_1_pane_2)

        self.Layout()

        self.Bind(wx.EVT_BUTTON, self.SelectFiles, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.ReadHeaders, self.button_3)
        self.Bind(wx.EVT_BUTTON, self.ExportSelected, self.button_2)
        # end wxGlade

    def SelectFiles(self, event):  # wxGlade: MainFrame.<event_handler>
        # data = []

        print(p.open_files())

        extension = self.text_ctrl_1.GetValue()
        extension = "*." + extension

        path = MyFileDialog(None, wildcard=extension)
        files = path.EventHandler.Paths
        sysno = []
        datum = []
        sysinfo = []
        sysno_f = []
        for file in files:
            found = False

            workbook = openpyxl.load_workbook(
                filename=file, data_only=True, read_only=True)

            ws = workbook['Daily Report']
            ws2 = workbook['Email Summary']

            status_no = ws2['R6'].value
            if status_no == 1:
                status = 'Mobilization'
            elif status_no == 2:
                status = 'System Assembly'
            elif status_no == 3:
                status = 'Testing'
            elif status_no == 4:
                status = 'Operational'
            elif status_no == 5:
                status = 'Troubleshooting'
            elif status_no == 6:
                status = 'Heli Tech'
            elif status_no == 7:
                status = 'Standby'
            elif status_no == 8:
                status = 'Admin/Safety'
            elif status_no == 9:
                status = 'Demobilization'

            try:
                comment = ws['A41'].value.replace(',', '')
            except:
                comment = ''

            client = ws['S4'].value.replace(',', '')
            system_conf = ws['CF2'].value
            datuminside = ws['BV3'].value.strftime('%Y/%m/%d')
            job_no = ws['BV4'].value
            try:
                location = ws['S5'].value.replace(',', '')
            except:
                location = ''
            chief = ws['M7'].value
            operator = ws['M8'].value
            pm = ws['M9'].value
            qc = ws['M10'].value
            newlines = ws['A15'].value
            reflights = ws['L15'].value
            dailytotal = ws['W15'].value
            qckm = ws['BP35'].value
            qcperc = ws['AS15'].value
            totalkm = ws['BD15'].value
            acctotal = ws['BO15'].value
            perccomp = ws['BZ15'].value
            tofly = ws['CK15'].value
            sysno = ws['CF1'].value

            workbook.close()

            print(f'{datuminside},{sysno},{system_conf},{job_no},{location},'
                  f'{client},{pm},{qc},{chief},{operator},{status},'
                  f'{newlines},{reflights},{dailytotal},{qckm},{totalkm},'
                  f'{acctotal},{perccomp},{tofly},{comment}')

            file_name = file.split('\\')[-1]
            for field in file_name.split():
                if (field.startswith('VTEM')) or (field.startswith('ZTEM')) or (field.startswith('FWZTEM')):
                    sysno_f.append(field)
                elif field.endswith('xlsm'):
                    datum.append(field.split('.')[0][-8:])
                # print(field)
            # print(file_name)
        for s in range(0, len(sysno_f)):
            sysinfo.append(sysno_f[s] + ' ' + datum[s])
        # sysno = [file.rsplit('\\')[-1].split(' ')[-2] for file in files]
        # datum = [file.rsplit('\\')[-1].split(' ')[-1].split('.')[0][-8:] for file in files]
        # print()


        self.check_list_box_1.SetItems(sysinfo)

        # print(p.open_files())

        event.Skip()

    def ReadHeaders(self, event):  # wxGlade: MainFrame.<event_handler>
        data = []
        headers = []

        extension = self.text_ctrl_1.GetValue()

        files = self.check_list_box_1.GetItems()

        for file_no in range(0, len(files)):

            try:
                f = open(files[file_no], "r")
                print(f"Reading file: {files[file_no]}")
                for line in f:
                    headers.append(line.rsplit(',')[0].rstrip())
                self.check_list_box_1.SetSelection(file_no)
                # self.check_list_box_1.SetCheckedItems(file_no)
            except:
                print("error in file {0}".format(files[file_no]))

        headers = list(set(headers))
        # headers = [header.rstrip() for header in headers]
        # headers = list(set(headers))
        headers = sorted(headers)
        self.check_list_box_2.SetItems(headers)

        print("Done reading headers")
        event.Skip()

    def ExportSelected(self, event):  # wxGlade: MainFrame.<event_handler>
        headers = self.check_list_box_2.GetCheckedStrings()

        path = MyNewFileDialog(None)
        output_file = path.EventHandler.Path
        print(output_file)

        print(self.check_list_box_1.GetItems())

        files = self.check_list_box_1.GetItems()

        with open(output_file, 'w') as out:
            for no, file in enumerate(files):
                print(f"Reading from: {file}")
                self.check_list_box_1.SetSelection(no)
                with open(file, 'r') as data_file:
                    for line in data_file:
                        for header in headers:
                            if line.startswith(header):
                                # print(line)
                                out.write(line)
        print(headers)
        print("Created file")
        event.Skip()

# end of class MainFrame

class MyFileDialog(wx.FileDialog):
    def __init__(self, *args, **kwds):
        # begin wxGlade: MyFileDialog.__init__
        kwds["style"] = kwds.get("style", 0) | wx.FD_OPEN | \
                        wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE
        wx.FileDialog.__init__(self, *args, **kwds)
        self.SetTitle("Select files")

        self.ShowModal()
# end of class MyFileDialog


class MyNewFileDialog(wx.FileDialog):
    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.FD_SAVE
        kwds['defaultFile'] = 'data.csv'
        wx.FileDialog.__init__(self, *args, **kwds)
        self.SetTitle("Enter filename")

        self.ShowModal()
# end of class MyFileDialog


class MyApp(wx.App):
    def OnInit(self):
        self.frame = MainFrame(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True

# end of class MyApp

if __name__ == "__main__":
    app = MyApp(0)
    app.MainLoop()